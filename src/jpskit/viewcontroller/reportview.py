"""pip uninstall pdfminer" and "pip uninstall pdfminer.six".
Then all clear, "pip install pdfminer.six".
https://www.bedjango.com/blog/how-generate-pdf-django-weasyprint/"""

from django.shortcuts import render, get_list_or_404, redirect, reverse
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse, Http404
from django.conf.urls.static import static
from django.conf import settings
from django.http import StreamingHttpResponse, HttpResponse, HttpResponseServerError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.humanize.templatetags.humanize import intcomma
import datetime
from PyPDF2 import PdfFileReader, PdfFileWriter, pdf
import io
from io import BytesIO
import os
import gspread
import json
from oauth2client.service_account import ServiceAccountCredentials
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from pdfjinja import PdfJinja
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from pdfjinja import PdfJinja
from tempfile import NamedTemporaryFile
from ..modelcontroller import document, dfnoperolehan, kontraktor, project, userprofile
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db.models import Count
from weasyprint import HTML, CSS
from weasyprint.fonts import FontConfiguration
from django.template.loader import get_template



@login_required(login_url='login')
def testexcel(request):
    
    wb = load_workbook('static_in_env/assets/excel/laporan.xlsx')
    for ws in wb.worksheets:
        print(ws.title)

    row_num = 7
    ws = wb.worksheets[0]
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    columns = [
      '1', 
     'B13-28209',
     'PENYELENGGARAAN TERUSAN, TALIAIR TANAH/KONKRIT SERTA PARIT DIKAWASAN RANTAU PANJANG, SKIM PENGAIRAN KOTA II, DAERAH KUALA MUDA, KEDAH DARUL AMAN.',
     '61,775.11',
     '(1) FEBAH ENTERPRISE\n(2)JPSKMSB(KU/KM)S/B/OM/101/2020\n(3)22-09-2020 / 12-11-2020\n(4) 26-08-2020 / 21-10-2020 ',
     '20,000.00',
     '61,775.11',
     '80,000.00',
     '-16,680.00',
     '80,000.00',
     '100%',
    ]
    columns2 = [
      '1', 
     'B13-28209',
     'PENYELENGGARAAN TERUSAN, TALIAIR TANAH/KONKRIT SERTA PARIT DIKAWASAN RANTAU PANJANG, SKIM PENGAIRAN KOTA II, DAERAH KUALA MUDA, KEDAH DARUL AMAN.',
     '61,775.11',
     '(1) FEBAH ENTERPRISE\n(2)JPSKMSB(KU/KM)S/B/OM/101/2020\n(3)22-09-2020 / 12-11-2020\n(4) 26-08-2020 / 21-10-2020 ',
     '20,000.00',
     '61,775.11',
     '80,000.00',
     '-16,680.00',
     '80,000.00',
     '100%',
    ]
 
 
   
    for col_num in range(len(columns)):
        ws.cell(row_num, col_num+1, columns[col_num]).border = thin_border
        ws.cell(row_num+1, col_num+1, columns2[col_num]).border = thin_border
            
    
    wb.save("render2.xlsx")
   

    return render(request, 'pages/someexcel.html')

@login_required(login_url='login')
def testexcel2(request):

    wb = load_workbook('static_in_env/assets/excel/laporan.xlsx')
    for ws in wb.worksheets:
        print(ws.title)

    row_num = 6
    ws = wb.worksheets[0]
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    data = document.MRKSatu.objects.all()


    for d in data:

        r1 = project.Projek.objects.filter(nosebuthargaid=d.projekbind.nosebuthargaid).first()
        r2 = document.kosprojek.objects.filter(projekbind=r1.id).first()

        columns = [""]
        columns1 = [d.mrksatunoinden]
        columns2 = [r1.tajukkerja]
        columns3 = [d.mrksatukosprojek]
        columns4 = ['(1)'+d.mrksatukontraktor.konNama+'\n(2)'+d.projekbind.nosebuthargaid.noperolehan+'\n(3)'+str(d.mrksatutarikhmula)+'/'+str(d.mrksatutarikhjangkasiap)+'\n(4)'+str(d.mrksatutarikhmula)+'/'+str(d.mrksatutarikhjangkasiap)+'']
        columns5 = [r1.peruntukansemasa]
        columns6 = [r2.kos_belanja ]
        columns7 = [r2.kos_tanggung]

        for col_num in range(len(columns)):   
            
            print(col_num)
            row_num +=1
            ws.cell(row_num, col_num+1, columns[col_num]).border = thin_border
            ws.cell(row_num, col_num+2, columns1[col_num]).border = thin_border
            ws.cell(row_num, col_num+3, columns2[col_num]).border = thin_border
            ws.cell(row_num, col_num+4, columns3[col_num]).border = thin_border
            ws.cell(row_num, col_num+5, columns4[col_num]).border = thin_border
            ws.cell(row_num, col_num+6, columns5[col_num]).border = thin_border
            ws.cell(row_num, col_num+7, columns6[col_num]).border = thin_border
            ws.cell(row_num, col_num+8, columns7[col_num]).border = thin_border
            ws.cell(row_num, col_num+9).border = thin_border
            ws.cell(row_num, col_num+10).border = thin_border
            ws.cell(row_num, col_num+11).border = thin_border
            wb.save("render2.xlsx")
        

    return render(request, 'pages/someexcel.html')

@login_required(login_url='login')
def exceltest(request):
     
    wb = load_workbook('static_in_env/assets/excel/bookpython.xlsx')
    for ws in wb.worksheets:
        print(ws.title)
    
    row_num = 4
    ws = wb.worksheets[0]
    columns = ['Username', 'Line 1\nLine 2\nLine 3', 'Last name', 'Email address', ]
    columns1 = ['Username', 'Line 1\nLine 2\nLine 3', 'Last name', 'Email address', ]
   
    for col_num in range(len(columns)):
        ws.cell(row_num, col_num+2, columns[col_num])
        ws['C'+ str(row_num+range)].alignment = Alignment(wrapText=True)
        ws.cell(row_num+1, col_num+2, columns1[col_num])
    
    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.cell(row_num, col_num+2, row[col_num])
    
    wb.save("chart.xlsx")


@login_required(login_url='login')
def pdfmrksatu(request, projekid):

    dataobject = document.MRKSatu.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/MRK-01.pdf')
    pdfout = pdfjinja(
        dict(
            sijilpkk=dataobject.mrksatukontraktor.sijilPPKNoPendaftaran,
            kontraktor=dataobject.mrksatukontraktor.konNama,
            nokontrak=dataobject.projekbind.nosebuthargaid,
            noinden=dataobject.mrksatunoinden,
            daereah=dataobject.mrksatukontraktor.konKawOperasi,
            negeri=dataobject.mrksatukontraktor.konNegeri,
            kosprojek=intcomma(dataobject.mrksatukosprojek),
            tarikhmula=dataobject.mrksatutarikhmula,
            tarikhjsiap=dataobject.mrksatutarikhjangkasiap,
            pegawai=dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
            jawatan=dataobject.projekbind.nosebuthargaid.pegawaiselia.userprofile.jawatan,
            tarikhlapo=dataobject.mrksatutarikhdaftar,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/MRK-01-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/MRK-01-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfmrkdua(request, projekid):

    dataobject = document.MRKDua.objects.filter(projekbind=projekid).first()

    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/MRK-02.pdf')
    pdfout = pdfjinja(
        dict(
            nopkk = dataobject.mrksatulink.mrksatukontraktor.sijilPPKNoPendaftaran,
            kontraktor = dataobject.mrksatulink.mrksatukontraktor.konNama,
            nokontrak = dataobject.projekbind.nosebuthargaid,
            noinden = dataobject.mrksatulink.mrksatunoinden,
            kosprojek = intcomma(dataobject.mrksatulink.mrksatukosprojek),
            tarikhmilik = dataobject.mrksatulink.mrksatutarikhmula,
            tarikhjangkasiap = dataobject.mrksatulink.mrksatutarikhjangkasiap,
            kemajuanprojek = dataobject.mrkduakerjajadual,
            kemajuankerjasebenar = dataobject.mrkduakerjasebenartarikh,
            sehingga = dataobject.mrkduakerjasebenar,
            nobayaran = dataobject.mrkduakemajuan,
            jumlahbayaran = intcomma(dataobject.mrkduabayarankemajuan),
            check1=dataobject.mrkduamodal,
            check2=dataobject.mkrduabahan,
            check3=dataobject.mrkduapekerja,
            check4=dataobject.mrkduatapak,
            check5=dataobject.mrkduacuaca,
            disebabkan = dataobject.mrkduadisebabkanoleh,
            lainlain = dataobject.mrkdualainlain,
            lanjutmasa = dataobject.mrkdualanjutmasa,
            lanjutdari = dataobject.mrkdualanjutdari,
            lanjuthingga = dataobject.mrkdualanjutsehingga,
            disebabkanoleh = dataobject.mrkduadisebabkan,
            lad = dataobject.mrkduaLAD,
            laddari = dataobject.mrkduaLADdari,
            ladhingga = dataobject.mrkduaLADSehingga,
            tarikhperakuan = dataobject.mrkduaperakuan,
            tarikhmansuh = dataobject.mrkduamansuh,

        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/MRK-02-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/MRK-02-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdflsk(request, projekid):

    dataobject = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
   

    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/Laporansiapkerja.pdf')
    pdfout = pdfjinja(
        dict(
            kontraktor = dataobject.lskmrksatulink.mrksatukontraktor.konNama,
            alamat1 = dataobject.lskmrksatulink.mrksatukontraktor.konAlamat,
            alamat2 = dataobject.lskmrksatulink.mrksatukontraktor.konAlamatExtS,
            poskod = dataobject.lskmrksatulink.mrksatukontraktor.konPoskod,
            bandar = dataobject.lskmrksatulink.mrksatukontraktor.konBandar,
            negeri = dataobject.lskmrksatulink.mrksatukontraktor.konNegeri,
            kerja = dataobject.projekbind.tajukkerja,
            inden = dataobject.lskmrksatulink.mrksatunoinden,
            sebutharga = dataobject.projekbind.nosebuthargaid,
            workmen = dataobject.lskperkeso,
            ins1 = dataobject.lsknoinsurancesatu,
            ins2 = dataobject.lsknoinsurancedua,
            peruntukan = dataobject.lskperuntukan,
            kosprojek = intcomma(dataobject.lskmrksatulink.mrksatukosprojek),
            hargasebenar = intcomma(dataobject.lskhargasebenar),
            tarikhmula = dataobject.lskmrksatulink.mrksatutarikhmula,
            tarikhtamat = dataobject.lskmrksatulink.mrksatutarikhjangkasiap,
            tarikhlanjutmasa = dataobject.lsklanjutmasa,
            tarikhsiapsem = dataobject.lskkerjasiap,
            laporan = dataobject.lsklaporan,
            tarikhakui = dataobject.lsktarikhperakui,
            penyelia = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
            jawatan = dataobject.projekbind.nosebuthargaid.pegawaiselia.userprofile.jawatan,
            kb = dataobject.lskketuabahagian,
            jawatankb = dataobject.lskjawatanketuabahagian,
            jurutera = dataobject.lskjuruteraj41,
            jawatanjurutera = dataobject.lskjawatanj41,
            jd = dataobject.lskjuruteradaerah,
            jawatanjd = dataobject.lskjawatanjuruteradaerah,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/Laporansiapkerja-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/Laporansiapkerja-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfmrktiga(request, projekid):

    dataobject = document.MRKTiga.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    mrkduaf = document.MRKDua.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/MRK03.pdf')
    pdfout = pdfjinja(
        dict(
            pkk = dataobject.marktigamrksatu.mrksatukontraktor.sijilPPKNoPendaftaran,
            kontraktor =dataobject.marktigamrksatu.mrksatukontraktor.konNama,
            sebutharga =dataobject.projekbind.nosebuthargaid,
            inden = dataobject.marktigamrksatu.mrksatunoinden,
            tajuk = dataobject.projekbind.tajukkerja,
            kosprojek = intcomma(dataobject.marktigamrksatu.mrksatukosprojek),
            kossebenar = intcomma(lskfetch.lskhargasebenar),
            tarikhmula = dataobject.marktigamrksatu.mrksatutarikhmula,
            tarikhakhir =lskfetch.lskkerjasiap,
            lanjuthingga = lskfetch.lsklanjutmasa,
            tarikhsiapsebenar =lskfetch.lskkerjasiap,
            lad =mrkduaf.mrkduaLAD,
            laddari =mrkduaf.mrkduaLADdari,
            ladhingga =mrkduaf.mrkduaLADSehingga,
            c1a = dataobject.mrktigabina,
            c1b = dataobject.mrktigabina,
            c1c = dataobject.mrktigabina,
            c1d = dataobject.mrktigabina,
            c2a = dataobject.mrktigatadbir,
            c2b = dataobject.mrktigatadbir,
            c2c = dataobject.mrktigatadbir,
            c2d = dataobject.mrktigatadbir,
            c3a = dataobject.mrktigakemajuan,
            c3b = dataobject.mrktigakemajuan,
            c3c = dataobject.mrktigakemajuan,
            c3d = dataobject.mrktigakemajuan,
            c4a = dataobject.mkrtigamutukerangka,
            c4b = dataobject.mkrtigamutukerangka,
            c4c = dataobject.mkrtigamutukerangka,
            c4d = dataobject.mkrtigamutukerangka,
            c5a = dataobject.mrktigamutukerja,
            c5b = dataobject.mrktigamutukerja,
            c5c = dataobject.mrktigamutukerja,
            c5d = dataobject.mrktigamutukerja,
            c6a = dataobject.mrktigamutukemasan,
            c6b = dataobject.mrktigamutukemasan,
            c6c = dataobject.mrktigamutukemasan,
            c6d = dataobject.mrktigamutukemasan,
            c7a = dataobject.mrktigamutukerjaluar,
            c7b = dataobject.mrktigamutukerjaluar,
            c7c = dataobject.mrktigamutukerjaluar,
            c7d = dataobject.mrktigamutukerjaluar,
            c8a = dataobject.mrktigapegawasan,
            c8b = dataobject.mrktigapegawasan,
            c8c = dataobject.mrktigapegawasan,
            c8d = dataobject.mrktigapegawasan,
            catat1 = dataobject.mrkcatat1,
            catat2 = dataobject.mrkcatat2,
            catat3 = dataobject.mrkcatat3,
            catat4 = dataobject.mrkcatat4,
            catat5 = dataobject.mrkcatat5,
            catat6 = dataobject.mrkcatat6,
            catat7 = dataobject.mrkcatat7,
            catat8 = dataobject.mrkcatat8,
            ulasan =dataobject.mrktigasokongan,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/MRK03-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/MRK03-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfpsksatu(request, projekid):

    dataobject = document.PSK.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    mrkduaf = document.MRKDua.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/PSK01.pdf')
    pdfout = pdfjinja(
        dict(

            sebutharga = dataobject.projekbind.nosebuthargaid,
            kontraktor = dataobject.pskmrksatulink.mrksatukontraktor.konNama,
            alamat1 = dataobject.pskmrksatulink.mrksatukontraktor.konAlamat,
            alamat2 = dataobject.pskmrksatulink.mrksatukontraktor.konAlamatExtS,
            poskod = dataobject.pskmrksatulink.mrksatukontraktor.konPoskod,
            bandar = dataobject.pskmrksatulink.mrksatukontraktor.konBandar,
            negeri = dataobject.pskmrksatulink.mrksatukontraktor.konNegeri,
            gred = dataobject.pskmrksatulink.mrksatugred,
            tajukkerja = dataobject.projekbind.tajukkerja,
            hargasebenar = intcomma(lskfetch.lskhargasebenar),
            tarikhsebenar = lskfetch.lskkerjasiap,
            tarikhambil = dataobject.psktarikhambilmilik,
            mulacacat = dataobject.psktarikhmulatanggug,
            akhircacat = dataobject.psktarikhtamattanggung,
            jurutere = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
  
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/PSK01-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/PSK01-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


@login_required(login_url='login')
def pdfpskdua(request, projekid):

    dataobject = document.PSK.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    mrkduaf = document.MRKDua.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/PKS02.pdf')
    pdfout = pdfjinja(
        dict(

            sebutharga = dataobject.projekbind.nosebuthargaid,
            kontraktor = dataobject.pskmrksatulink.mrksatukontraktor.konNama,
            alamat1 = dataobject.pskmrksatulink.mrksatukontraktor.konAlamat,
            alamat2 = dataobject.pskmrksatulink.mrksatukontraktor.konAlamatExtS,
            poskod = dataobject.pskmrksatulink.mrksatukontraktor.konPoskod,
            bandar = dataobject.pskmrksatulink.mrksatukontraktor.konBandar,
            negeri = dataobject.pskmrksatulink.mrksatukontraktor.konNegeri,
            gred = dataobject.pskmrksatulink.mrksatugred,
            tajukkerja = dataobject.projekbind.tajukkerja,
            hargasebenar = intcomma(lskfetch.lskhargasebenar),
            tarikhsebenar = lskfetch.lskkerjasiap,
            tarikhambil = dataobject.psktarikhambilmilik,
            mulacacat = dataobject.psktarikhmulatanggug,
            akhircacat = dataobject.psktarikhtamattanggung,
            jurutere = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name
  
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/PKS02-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/PKS02-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def ssemak(request, projekid):

    dataobject = document.SenaraiSemakan.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    mrkduaf = document.MRKDua.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/SS.pdf')
    pdfout = pdfjinja(
        dict(
            check1 = dataobject.ssinden,
            check2 = dataobject.sslsk,
            check3 = dataobject.ssti,
            check4 = dataobject.sssebutharga,
            check5 = dataobject.sspt,
            check6 = dataobject.ssjs,
            check7 = dataobject.sskts,
            check8 = dataobject.ssds,
            check9 = dataobject.ssplm,
            check10 = dataobject.ssab,
            check11 = dataobject.sscidb,
            check12 = dataobject.sspkk,
            check13 = dataobject.ssssm,
            check14 = dataobject.sskk,
            check15 = dataobject.ssinsurance,
            check16 = dataobject.ssgambar,
            pegawai = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
            tarikh = dataobject.sstarikh
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/SS-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/SS-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfpsmk(request, projekid):

    dataobject = document.PSMK.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/PSMK.pdf')
    pdfout = pdfjinja(
        dict(

            kontrator = dataobject.psmkmrksatulink.mrksatukontraktor.konNama,
            alamat1 = dataobject.psmkmrksatulink.mrksatukontraktor.konAlamat,
            alamat2 = dataobject.psmkmrksatulink.mrksatukontraktor.konAlamatExtS,
            poskod = dataobject.psmkmrksatulink.mrksatukontraktor.konPoskod,
            bandar = dataobject.psmkmrksatulink.mrksatukontraktor.konBandar,
            negeri = dataobject.psmkmrksatulink.mrksatukontraktor.konNegeri,
            gred = dataobject.psmkmrksatulink.mrksatugred,
            sebutharga = dataobject.projekbind.nosebuthargaid,
            tajuk = dataobject.projekbind.tajukkerja,
            tarikhgagal = lskfetch.lskkerjasiap,
            g1 = dataobject.psmknojaminanbanka,
            hargaa = intcomma(dataobject.psmkhargajaminana),
            bakia = intcomma(dataobject.psmkbakiwangjaminana),
            g2 = dataobject.psmknojaminanbankab,
            hargab = intcomma(dataobject.psmkhargajaminanb),
            wangb = intcomma(dataobject.psmkbakiwangjaminanb),
            kosbon = intcomma(dataobject.psmkkosbon),
            bakikosbon = intcomma(dataobject.psmkbakikos),
            pegawai = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
            jawatan = dataobject.projekbind.nosebuthargaid.pegawaiselia.userprofile.jawatan,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/PSMK-'+str(idstr)+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/PSMK-'+str(idstr)+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfjb(request, projekid):

    dataobject = document.SuratPJaminanbank.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/jb.pdf')
    pdfout = pdfjinja(
        dict(
            namabank = dataobject.namabank,
            alamatbank = dataobject.alamatbank,
            akaunbank = dataobject.rujukanbank,
            kontraktor = dataobject.jbankmrksatulink.mrksatukontraktor.konNama,
            alamat = dataobject.alamatpemborongsurat,
            tarikha = psk.psktarikhmulatanggug,
            tarikhb = psk.psktarikhtamattanggung,
            jurutera = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
            jawatan = dataobject.projekbind.nosebuthargaid.pegawaiselia.userprofile.jawatan,

 
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/jb-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/jb-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfppwjp(request, projekid):

    dataobject = document.Perakuanpwjp.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/ppwjp.pdf')
    pdfout = pdfjinja(
        dict(
            namarujukan = dataobject.namarujukan,
            alamtrujukan = dataobject.alamatrujukan,
            kontraktor = dataobject.wjpmrksatulink.mrksatukontraktor.konNama,
            projek = dataobject.projekbind.tajukkerja,
            rm = dataobject.koswjp,
            jurutera = dataobject.projekbind.nosebuthargaid.pegawaiselia.first_name,
            jawatan = dataobject.projekbind.nosebuthargaid.pegawaiselia.userprofile.jawatan,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/ppwjp-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/ppwjp-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


@login_required(login_url='login')
def pdfsmrksatu(request, projekid):

    dataobject = document.SuratMRK.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/spkk1.pdf')
    pdfout = pdfjinja(
        dict(

            namarujukan = dataobject.smrknamarujukan,
            alamatrujukan =dataobject.smkralamatrujukan,
            jenisborang = dataobject.smrkjenisborang,
            kontraktor = dataobject.smrkmrksatulink.mrksatukontraktor.konNama,
            nopkk = dataobject.smrkmrksatulink.mrksatukontraktor.sijilPPKNoPendaftaran,
            norujukan = dataobject.smrkrujukantuan,
            pegawai = dataobject.smrkpegawai,
            jawatan = dataobject.smrkjawatan,
    
        ))


    pdfout.write(open('static_in_env/assets/pdf/outputpdf/spkk1-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/spkk1-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


@login_required(login_url='login')
def pdfsmrkdua(request, projekid):

    dataobject = document.SuratMRK.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/spkk2.pdf')
    pdfout = pdfjinja(
        dict(
            namarujukan = dataobject.smrknamarujukan,
            alamatrujukan =dataobject.smkralamatrujukan,
            jenisborang = dataobject.smrkjenisborang,
            kontraktor = dataobject.smrkmrksatulink.mrksatukontraktor.konNama,
            nopkk = dataobject.smrkmrksatulink.mrksatukontraktor.sijilPPKNoPendaftaran,
            norujukan = dataobject.smrkrujukantuan,
            pegawai = dataobject.smrkpegawai,
            jawatan = dataobject.smrkjawatan,
            tarikh = dataobject.smrktarikh,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/spkk2-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/spkk2-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


@login_required(login_url='login')
def pdfskhas01(request, projekid):

    dataobject = document.SuratKhas.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/skhas01.pdf')
    pdfout = pdfjinja(
        dict(
            rujukantuan = dataobject.khasrujukantuan,
            rujukan = dataobject.projekbind,
            namarujukan = dataobject.khasnamarujukan,
            alamatrujukan =dataobject.khasalamatrujukan,
            gred = dataobject.khasmrksatulink.mrksatugred,
            kategori = dataobject.khasmrksatulink.mrksatukategori,
            khusus  = dataobject.khasmrksatulink.mrksatupengkhususan,
            kontraktor = dataobject.khasmrksatulink.mrksatukontraktor.konNama,
            projek = dataobject.projekbind.tajukkerja,
            jurutera = dataobject.khaspegawai,
            jawatan = dataobject.khasjawatan,

        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/skhas01-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/skhas01-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


@login_required(login_url='login')
def pdfskhas02(request, projekid):

    dataobject = document.SuratKhas.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/skhas02.pdf')
    pdfout = pdfjinja(
        dict(
            rujukantuan = dataobject.khasrujukantuan,
            rujukan = dataobject.projekbind,
            namarujukan = dataobject.khasnamarujukan,
            alamatrujukan =dataobject.khasalamatrujukan,
            gred = dataobject.khasmrksatulink.mrksatugred,
            kategori = dataobject.khasmrksatulink.mrksatukategori,
            khusus  = dataobject.khasmrksatulink.mrksatupengkhususan,
            kontraktor = dataobject.khasmrksatulink.mrksatukontraktor.konNama,
            projek = dataobject.projekbind.tajukkerja,
            jurutera = dataobject.khaspegawai,
            jawatan = dataobject.khasjawatan,
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/skhas02-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/skhas02-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


@login_required(login_url='login')
def pdfpwjp01(request, projekid):

    dataobject = document.SuratPelepasanBon.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/spwjp.pdf')
    pdfout = pdfjinja(
        dict(
          
            rujukan = dataobject.projekbind,
            kepada = dataobject.bonkepada,
            alamatkepada =dataobject.bonalamatsatu,
            kontraktor = dataobject.bonmrksatulink.mrksatukontraktor.konNama,
            kepada2 = dataobject.bonmelalui,
            alamatkepada2 = dataobject.bonalamatdua,
            jurutera = dataobject.bonpegawai,
            jawatan = dataobject.bonjawatan,
            kosbon = intcomma(dataobject.bonwangjaminan),
      
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/spwjp-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/spwjp-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )

@login_required(login_url='login')
def pdfpwjp02(request, projekid):

    dataobject = document.SuratPelepasanBon.objects.filter(projekbind=projekid).first()
    lskfetch = document.Laporansiapkerja.objects.filter(projekbind=projekid).first()
    psk  = document.PSK.objects.filter(projekbind=projekid).first()
 
    idstr = str(dataobject.projekbind.id)
    print(idstr)
 
    pdfjinja = PdfJinja('static_in_env/assets/pdf/spwjp02.pdf')
    pdfout = pdfjinja(
        dict(
            
            rujukan = dataobject.projekbind,
            kepada = dataobject.bonkepada,
            alamatkepada =dataobject.bonalamatsatu,
            kontraktor = dataobject.bonmrksatulink.mrksatukontraktor.konNama,
            kepada2 = dataobject.bonmelalui,
            alamatkepada2 = dataobject.bonalamatdua,
            jurutera = dataobject.bonpegawai,
            jawatan = dataobject.bonjawatan,
            kosbon = intcomma(dataobject.bonwangjaminan),
      
        ))

 
    pdfout.write(open('static_in_env/assets/pdf/outputpdf/spwjp02-'+idstr+'.pdf', 'wb'))
    try:
        return FileResponse(open('static_in_env/assets/pdf/outputpdf/spwjp02-'+idstr+'.pdf', 'rb'), content_type='application/pdf')
    except FileNotFoundError:
        raise Http404()

    return render(request, 'pages/printtest.html' )


##filter the report
@login_required(login_url='login')
def report_by_filter(request):

    kaedah = request.GET.get('kaedah')
    timecurrent = datetime.date.today().strftime('%d/%m/%Y')
    print(kaedah)

    qs = document.MRKSatu.objects.filter(projekbind__nosebuthargaid__tarikh__year=2021).values(
        'projekbind__kodvot__no',
        'mrksatutarikhmula',
        'mrksatutarikhmula',
        'mrksatukosprojek',
        'projekbind__tajukkerja',
        'projekbind__peruntukansemasa',
        'projekbind__nosebuthargaid__noperolehan',
        'mrksatukontraktor__konNama',
        'bindone__mrkduaLADdari',
        'bindone__mrkduaLADSehingga',
        'projekbind__kosprojek__kos_belanja',
        'projekbind__kosprojek__kos_tanggung',
    
    )
    print(qs.query)
    

    context = {
        'qs':qs,
        'kd':project.KDvot.objects.all()
    }
  

    print(qs.query)
    
    return render(request, 'pages/laporan_filter.html',context)


def pdfhtmlgenerator(request):
    html_template = get_template('htmlprint/kontraktor.html').render()
    print(html_template)
    pdfgenerate = HTML(string=html_template).write_pdf(stylesheets=["https://fonts.googleapis.com/css?family=Raleway:400,600&display=swap"])
    trigger = HttpResponse(pdfgenerate, content_type='application/pdf' )
    trigger['Content-Disposition'] = 'filename="home_page.pdf"'
    return trigger
    







    
    
    
    
    
    
    
    

    
